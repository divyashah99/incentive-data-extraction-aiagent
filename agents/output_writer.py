"""
OutputWriter — writes validated IncentiveRecord objects to the Phase 0 handoff
package (spec §6.4):

  1. extracted_tampa_incentives.csv  — Layer 1, clean rows only (spec §4 cols)
  2. quarantine.csv                  — failed rows + quarantine_reason column
  3. program_geo.csv                 — Layer 2 search index (spec §5.2)
  4. extracted_tampa_incentives.xlsx — multi-sheet workbook for human review

The CSV/quarantine split replaces the old in-row ``review_needed`` flag.
"""
import csv
import hashlib
import logging
import os
import re
from pathlib import Path

from validators.schema_validator import IncentiveRecord, CSV_COLUMNS
from validators.zip_enricher import load_crosswalk_rows

logger = logging.getLogger(__name__)

# Excel sheet names are limited to 31 chars and forbid these characters
_INVALID_SHEET_CHARS = re.compile(r"[\\/?*\[\]:]")

# program_geo column order (spec §5.2)
PROGRAM_GEO_COLUMNS = ["program_id", "geo_type", "geo_value"]


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Make a sheet name Excel-safe (≤31 chars, no forbidden chars, unique)."""
    clean = _INVALID_SHEET_CHARS.sub(" ", name).strip()[:31] or "Sheet"
    if clean not in used:
        used.add(clean)
        return clean
    for i in range(2, 100):
        suffix = f" ({i})"
        candidate = clean[: 31 - len(suffix)] + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
    used.add(clean)
    return clean


def _program_id(record: IncentiveRecord) -> str:
    """
    Stable placeholder id for the handoff (spec §6.4). The platform lead
    assigns the real FK on load; this lets program_geo rows reference their
    parent Layer 1 row before Supabase ids exist.
    """
    key = f"{(record.program_name or '').strip().lower()}|{(record.program_links or '').strip().lower()}"
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]


def _split_zip_field(zip_value: str | None) -> list[str]:
    if not zip_value:
        return []
    parts = re.split(r"[;,]", zip_value)
    return [p.strip() for p in parts if p.strip()]


class OutputWriter:

    def __init__(self, output_path: str = "output/extracted_tampa_incentives.csv"):
        self.output_path = output_path
        self.output_dir = os.path.dirname(output_path) or "."

    # ── Main Layer 1 CSV (clean rows only) ───────────────────────────────────

    def write(self, records: list[IncentiveRecord]) -> int:
        """Write clean records to the main CSV (overwrite)."""
        os.makedirs(self.output_dir, exist_ok=True)

        with open(self.output_path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            for record in records:
                writer.writerow(record.to_csv_row())

        logger.info("Wrote %d records to %s", len(records), self.output_path)
        return len(records)

    def append(self, records: list[IncentiveRecord]) -> int:
        """Append records to existing CSV (creates file if not present)."""
        os.makedirs(self.output_dir, exist_ok=True)
        file_exists = os.path.exists(self.output_path)

        with open(self.output_path, "a", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS, extrasaction="ignore")
            if not file_exists:
                writer.writeheader()
            for record in records:
                writer.writerow(record.to_csv_row())

        logger.info("Appended %d records to %s", len(records), self.output_path)
        return len(records)

    # ── Quarantine file (spec §6.4 Phase 0 handoff) ──────────────────────────

    def write_quarantine(
        self,
        records: list[IncentiveRecord],
        path: str | None = None,
    ) -> int:
        """
        Write quarantined records to quarantine.csv with the 13 spec columns
        plus a trailing ``quarantine_reason`` column (semicolon-joined reasons).
        Replaces the deprecated in-row ``review_needed`` flag.
        """
        qpath = path or os.path.join(self.output_dir, "quarantine.csv")
        os.makedirs(os.path.dirname(qpath) or ".", exist_ok=True)
        fieldnames = CSV_COLUMNS + ["quarantine_reason"]

        with open(qpath, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for record in records:
                row = record.to_csv_row()
                row["quarantine_reason"] = "; ".join(record.quarantine_reasons)
                writer.writerow(row)

        logger.info("Wrote %d quarantined records to %s", len(records), qpath)
        return len(records)

    # ── program_geo (Layer 2, spec §5) ────────────────────────────────────────

    def write_program_geo(
        self,
        records: list[IncentiveRecord],
        path: str | None = None,
    ) -> int:
        """
        Build the Layer 2 search index from validated (clean) records.
        Spec §5.1: every active program needs at least one program_geo row.
        Expansion rules: spec §5.3.
        """
        gpath = path or os.path.join(self.output_dir, "program_geo.csv")
        os.makedirs(os.path.dirname(gpath) or ".", exist_ok=True)

        crosswalk = load_crosswalk_rows()
        tampa_zips = sorted({r["zip"] for r in crosswalk if r.get("city", "").lower() == "tampa"})
        county_zips = sorted({r["zip"] for r in crosswalk if r.get("zip")})

        seen: set[tuple[str, str, str]] = set()
        geo_rows: list[dict] = []

        def add(pid: str, gtype: str, gvalue: str) -> None:
            key = (pid, gtype, gvalue)
            if not gvalue or key in seen:
                return
            seen.add(key)
            geo_rows.append({"program_id": pid, "geo_type": gtype, "geo_value": gvalue})

        for record in records:
            pid = _program_id(record)
            initial_count = len(geo_rows)

            zips_from_record = _split_zip_field(record.zip_code)
            city = (record.city or "").strip()
            state = (record.state or "").strip()

            # 1. Explicit ZIPs on the row → one zip row each
            for z in zips_from_record:
                add(pid, "zip", z)

            # 2. City-level scoping
            if city.lower() == "tampa":
                add(pid, "city", "Tampa")
                if not zips_from_record:
                    for z in tampa_zips:
                        add(pid, "zip", z)
            elif city.lower() in ("hillsborough county", "hillsborough"):
                add(pid, "county", "Hillsborough")
                if not zips_from_record:
                    for z in county_zips:
                        add(pid, "zip", z)
            elif city:
                add(pid, "city", city)

            # 3. State-level scoping (spec §5.3)
            if state.lower() == "florida" and not city:
                add(pid, "state", "Florida")
            elif state.lower() in ("all", "federal", "nationwide", "us"):
                add(pid, "nationwide", "US")

            # 4. Spec §5.1 guarantee — every program must have at least one row
            if len(geo_rows) == initial_count:
                # Fall back to whatever scoping we can derive; otherwise mark
                # nationwide so the row still appears in search instead of
                # being silently invisible. The platform reviewer can refine.
                add(pid, "nationwide", "US")

        with open(gpath, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=PROGRAM_GEO_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            for row in geo_rows:
                writer.writerow(row)

        logger.info(
            "Wrote %d program_geo rows for %d programs to %s",
            len(geo_rows), len(records), gpath,
        )
        return len(geo_rows)

    # ── Multi-sheet XLSX (review aid) ────────────────────────────────────────

    def write_xlsx(
        self,
        records_by_source: dict[str, list[IncentiveRecord]],
        combined: list[IncentiveRecord],
        path: str | None = None,
    ) -> int:
        """
        Write a multi-sheet Excel workbook.
          • "All Records" sheet first
          • One sheet per source (named after the source's display name or id)
        Quarantined rows are highlighted yellow for at-a-glance review.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        xlsx_path = path or self.output_path.replace(".csv", ".xlsx")
        os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        flag_fill = PatternFill("solid", fgColor="FFF2CC")  # soft yellow
        wrap = Alignment(wrap_text=True, vertical="top")

        # XLSX header includes the off-schema quarantine_reason so reviewers
        # can see why a row was flagged. The main CSV stays at 13 cols.
        xlsx_columns = CSV_COLUMNS + ["quarantine_reason"]

        def _add_sheet(sheet_name: str, records: list[IncentiveRecord]) -> None:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(xlsx_columns)
            for col_idx in range(1, len(xlsx_columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for record in records:
                row = record.to_csv_row()
                row_values = [row.get(c, "") for c in CSV_COLUMNS]
                row_values.append("; ".join(record.quarantine_reasons))
                ws.append(row_values)
                r = ws.max_row
                if record.is_quarantined:
                    for col_idx in range(1, len(xlsx_columns) + 1):
                        ws.cell(row=r, column=col_idx).fill = flag_fill
                for col_idx in range(1, len(xlsx_columns) + 1):
                    ws.cell(row=r, column=col_idx).alignment = wrap

            widths = {
                "program_name": 40, "state": 10, "city": 18, "zip_code": 16,
                "incentive_type": 18, "service_category": 24, "property_type": 22,
                "description": 60, "eligibility_criteria": 45,
                "incentive_amount": 28, "valid_until": 14, "updated_at": 14,
                "program_links": 40, "quarantine_reason": 30,
            }
            for i, col in enumerate(xlsx_columns, start=1):
                ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 20)
            ws.freeze_panes = "A2"

        used_names: set[str] = set()
        all_sheet = _safe_sheet_name("All Records", used_names)
        _add_sheet(all_sheet, combined)

        per_source_total = 0
        for source_id in sorted(records_by_source.keys()):
            recs = records_by_source[source_id]
            if not recs:
                continue
            sheet_name = _safe_sheet_name(source_id, used_names)
            _add_sheet(sheet_name, recs)
            per_source_total += len(recs)

        wb.save(xlsx_path)
        logger.info(
            "Wrote XLSX: %s (%d combined records, %d sheets per source)",
            xlsx_path, len(combined), len(records_by_source),
        )
        return per_source_total

    # ── Dedup ─────────────────────────────────────────────────────────────────

    def deduplicate(self, records: list[IncentiveRecord]) -> list[IncentiveRecord]:
        """
        Remove duplicates keyed on (program_name, incentive_type) — case-insensitive.
        When duplicates exist, prefer the record with no quarantine reasons.
        """
        seen: dict[tuple, IncentiveRecord] = {}

        for record in records:
            key = (
                (record.program_name or "").lower().strip(),
                (record.incentive_type or "").lower().strip(),
            )
            if key not in seen:
                seen[key] = record
            else:
                existing = seen[key]
                if (not record.is_quarantined) and existing.is_quarantined:
                    seen[key] = record

        deduped = list(seen.values())
        dropped = len(records) - len(deduped)
        if dropped:
            logger.info("Deduplication removed %d duplicate(s)", dropped)
        return deduped
